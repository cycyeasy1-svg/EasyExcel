# Generates M1 import-layer fixtures:
#   theme-colors.xlsx  - theme color fonts/fills with tints (Accent1-6)
#   numfmt-exotic.xlsx - exotic number formats that must pass through verbatim
#   richtext.xlsx      - in-cell rich text runs
#   borders.xlsx       - all 13 OOXML border styles
#   freeze-merge.xlsx  - frozen panes + merged ranges + row/col sizes
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

OUT = "test/fixtures"


def theme_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Theme"
    # theme indices per OOXML color.theme attr: 4..9 = accent1..6
    for i, theme in enumerate(range(4, 10), start=1):
        ws.cell(row=i, column=1, value=f"accent{theme - 3} font").font = Font(
            color=Color(theme=theme, type="theme")
        )
        c = ws.cell(row=i, column=2, value=f"accent{theme - 3} fill")
        c.fill = PatternFill(
            patternType="solid", fgColor=Color(theme=theme, type="theme")
        )
    # tint variants on accent1
    for i, tint in enumerate([0.4, -0.25], start=8):
        ws.cell(row=i, column=1, value=f"accent1 tint {tint}").font = Font(
            color=Color(theme=4, tint=tint, type="theme")
        )
    # indexed color
    ws.cell(row=11, column=1, value="indexed 17 (green)").font = Font(
        color=Color(indexed=17, type="indexed")
    )
    # explicit white fill (must be preserved, legacy reader drops it)
    c = ws.cell(row=12, column=1, value="white fill")
    c.fill = PatternFill(patternType="solid", fgColor="FFFFFFFF")
    wb.save(f"{OUT}/theme-colors.xlsx")


def numfmt_exotic():
    wb = Workbook()
    ws = wb.active
    ws.title = "NumFmt"
    cases = [
        (1234.5678, '#,##0.00'),
        (0.4567, '0.00%'),
        (1234.5, '¥#,##0.00;[Red]-¥#,##0.00'),
        (0.5, '# ?/?'),
        (12345678, '0.00E+00'),
        (3.14159, '0.00"m";[Red]-0.00"m"'),
        (45123, 'yyyy"年"m"月"d"日"'),
        (0.75, '[h]:mm:ss'),
        (42, '000000'),
        (1234, '[$-409]#,##0.00;[Red]([$-409]#,##0.00)'),
    ]
    for i, (v, fmt) in enumerate(cases, start=1):
        c = ws.cell(row=i, column=1, value=v)
        c.number_format = fmt
        ws.cell(row=i, column=2, value=fmt)
    wb.save(f"{OUT}/numfmt-exotic.xlsx")


def richtext():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rich"
    rt = CellRichText(
        TextBlock(InlineFont(b=True, color="FFFF0000"), "Bold red"),
        " plain ",
        TextBlock(InlineFont(i=True, sz=16), "big italic"),
    )
    ws["A1"] = rt
    ws["A2"] = "plain text"
    wb.save(f"{OUT}/richtext.xlsx")


def borders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Borders"
    styles = [
        "thin", "hair", "dotted", "dashed", "dashDot", "dashDotDot", "double",
        "medium", "mediumDashed", "mediumDashDot", "mediumDashDotDot",
        "slantDashDot", "thick",
    ]
    for i, s in enumerate(styles):
        row, col = i // 4 * 2 + 2, (i % 4) * 2 + 2
        c = ws.cell(row=row, column=col, value=s)
        side = Side(style=s, color="FF3366CC")
        c.border = Border(top=side, bottom=side, left=side, right=side)
    wb.save(f"{OUT}/borders.xlsx")


def freeze_merge():
    wb = Workbook()
    ws = wb.active
    ws.title = "FreezeMerge"
    ws.freeze_panes = "C3"  # 2 rows + 2 cols frozen
    ws.merge_cells("A1:B2")
    ws["A1"] = "merged"
    ws.merge_cells("D4:F4")
    ws["D4"] = "wide merge"
    for r in range(1, 12):
        ws.cell(row=r, column=3, value=r * 100)
    ws.row_dimensions[5].height = 40
    ws.column_dimensions[get_column_letter(4)].width = 30
    ws2 = wb.create_sheet("Second")
    ws2["A1"] = "second sheet"
    wb.save(f"{OUT}/freeze-merge.xlsx")


theme_colors()
numfmt_exotic()
richtext()
borders()
freeze_merge()
print("M1 fixtures written")
