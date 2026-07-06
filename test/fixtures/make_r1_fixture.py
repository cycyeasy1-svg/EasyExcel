# Generates test/fixtures/r1-features.xlsx containing features that must survive
# an ExcelJS load->write round trip: chart, comment, defined name, conditional
# formatting, rich text is added separately. Pivot/VBA can't be authored by
# openpyxl from scratch; chart is the representative "unmodeled part" probe.
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName

def base_workbook(with_chart: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Month", "Sales"])
    for i, (m, v) in enumerate([("Jan", 10), ("Feb", 25), ("Mar", 18), ("Apr", 32)], start=2):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)

    if with_chart:
        chart = BarChart()
        chart.title = "Sales by Month"
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D2")

    ws["A1"].comment = Comment("This is a comment on A1", "R1 Fixture")

    wb.defined_names["SalesRange"] = DefinedName("SalesRange", attr_text="Data!$B$2:$B$5")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add("B2:B5", CellIsRule(operator="greaterThan", formula=["20"], fill=red_fill))

    ws2 = wb.create_sheet("Untouched")
    ws2["A1"] = "this sheet is never edited"
    ws2["B1"] = 3.14159
    ws2["B1"].number_format = '0.00"m" ;[Red]-0.00"m"'
    return wb


base_workbook(with_chart=False).save("test/fixtures/r1-features.xlsx")
base_workbook(with_chart=True).save("test/fixtures/r1-chart.xlsx")
print("wrote r1-features.xlsx (no chart) + r1-chart.xlsx (with chart)")
