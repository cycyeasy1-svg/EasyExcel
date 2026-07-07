# M3 fixture: data validation (list + whole-between) + protected sheet +
# floating image, for feature-parity import/export tests.
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PilImage
import io

wb = Workbook()
ws = wb.active
ws.title = "Features"

ws["A1"] = "pick one"
dv_list = DataValidation(type="list", formula1='"red,green,blue"', allow_blank=True)
dv_list.add("B1:B5")
ws.add_data_validation(dv_list)

ws["A2"] = "1-100"
dv_num = DataValidation(type="whole", operator="between", formula1="1", formula2="100", allow_blank=False)
dv_num.add("C1:C5")
ws.add_data_validation(dv_num)

# floating image (8x8 red PNG)
img_bytes = io.BytesIO()
PilImage.new("RGB", (8, 8), (255, 0, 0)).save(img_bytes, format="PNG")
img_bytes.seek(0)
img = XlImage(img_bytes)
ws.add_image(img, "E3")

ws2 = wb.create_sheet("Locked")
ws2["A1"] = "protected sheet"
ws2["B1"] = "unlocked cell"
ws2["B1"].protection = ws2["B1"].protection.copy(locked=False)
ws2.protection.sheet = True

wb.save("test/fixtures/m3-features.xlsx")
print("wrote m3-features.xlsx")
